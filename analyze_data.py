#!/usr/bin/env python3
"""
更深入的数据探索 - 查看不同SKU和更多评论样本
"""
from openpyxl import load_workbook
from collections import Counter

excel_path = "/Users/AI 项目/Vibe Coding/ProductReviewAnalysis/电商产品评价-示例数据.xlsx"
wb = load_workbook(excel_path, data_only=True)
ws = wb.active

# 收集所有数据
data = []
for row_idx in range(2, ws.max_row + 1):  # 跳过表头
    row_data = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        value = ws.cell(row_idx, col_idx).value
        row_data[header] = value
    data.append(row_data)

print("=" * 80)
print("SKU 分布统计")
print("=" * 80)
sku_counter = Counter(item.get("SKU") for item in data)
for sku, count in sku_counter.most_common():
    print(f"  {sku}: {count} 条")

print("\n" + "=" * 80)
print("随机评论样本（含追评）")
print("=" * 80)

# 找一些有追评的数据
with_review = [item for item in data if item.get("追评内容")]
print(f"有追评的数量: {len(with_review)}\n")

# 显示10条不同类型的评论样本
sample_indices = [10, 50, 100, 200, 300, 400, 500, 600, 700, 800]
for i, idx in enumerate(sample_indices, 1):
    if idx < len(data):
        item = data[idx]
        print(f"【样本 {i}】SKU: {item.get('SKU')}")
        print(f"  初评: {item.get('初评内容')}")
        if item.get('追评内容'):
            print(f"  追评: {item.get('追评内容')}")
        print()
