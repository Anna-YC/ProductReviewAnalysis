#!/usr/bin/env python3
"""
探索Excel数据结构
"""
from openpyxl import load_workbook

excel_path = "/Users/AI 项目/Vibe Coding/ProductReviewAnalysis/电商产品评价-示例数据.xlsx"

wb = load_workbook(excel_path, data_only=True)
ws = wb.active

print("=" * 60)
print("Excel 文件结构分析")
print("=" * 60)
print(f"工作表名称: {wb.sheetnames}")
print(f"活动工作表: {ws.title}")
print(f"数据范围: {ws.dimensions}")
print()

print("=" * 60)
print("表头（前5列）")
print("=" * 60)
for idx, cell in enumerate(ws[1], 1):
    if idx > 10:  # 只显示前10列
        break
    print(f"  列{idx}: {cell.value}")
print()

print("=" * 60)
print("数据样本（前3行）")
print("=" * 60)
for row_idx in range(1, min(4, ws.max_row + 1)):
    print(f"\n--- 第 {row_idx} 行 ---")
    for col_idx in range(1, min(8, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        print(f"  列{col_idx}: {cell.value}")

print()
print("=" * 60)
print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")
print("=" * 60)
